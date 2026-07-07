"""
PP-OCRv6 Local Studio — Professional OCR Workbench
启动: python3 webapp/server.py
访问: http://localhost:8766
"""
import json, time, io, base64, sqlite3, uuid, os, csv, html, re, tempfile, subprocess
from collections import Counter
from contextlib import asynccontextmanager
from pathlib import Path
from datetime import datetime
from urllib.parse import quote
from html.parser import HTMLParser
import math
import numpy as np
import yaml
from PIL import Image, ImageDraw
import onnxruntime as ort
from fastapi import FastAPI, File, UploadFile, HTTPException, Query, Body, Form
from fastapi.responses import HTMLResponse, JSONResponse, FileResponse, PlainTextResponse, Response
from fastapi.staticfiles import StaticFiles
import uvicorn

try:
    import requests
except Exception:
    requests = None

try:
    import pypdfium2 as pdfium
except Exception:
    pdfium = None

try:
    import cv2
except Exception:
    cv2 = None

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except Exception:
    Workbook = None
    load_workbook = None

# ── 路径 ─────────────────────────────────────────────────────────────────────
ROOT      = Path(__file__).parent.parent
ONNX      = ROOT / "ppocrv6_onnx"
MODEL_VARIANTS = {
    "tiny": {
        "label": "PP-OCRv6 Tiny",
        "path": ROOT / "ppocrv6_onnx",
        "official_size": "1.5M",
        "use_case": "浏览器 / 极轻端侧",
    },
    "small": {
        "label": "PP-OCRv6 Small",
        "path": ROOT / "ppocrv6_small_onnx",
        "official_size": "7.7M",
        "use_case": "移动端 / 本地应用",
    },
    "medium": {
        "label": "PP-OCRv6 Medium",
        "path": ROOT / "ppocrv6_medium_onnx",
        "official_size": "34.5M",
        "use_case": "服务器 / 高精度本地",
    },
}
STATIC    = Path(__file__).parent / "static"
DATA      = Path(__file__).parent / "data"
UPLOADS   = DATA / "uploads"
ANNOTATED = DATA / "annotated"
THUMBS    = DATA / "thumbs"
TASK_PAGES = DATA / "task_pages"
DB_PATH   = DATA / "history.db"
CFG_PATH  = DATA / "config.json"
SECRET_BLOB_PATH = DATA / "secrets.blob"

for d in (UPLOADS, ANNOTATED, THUMBS, TASK_PAGES):
    d.mkdir(parents=True, exist_ok=True)

MAX_UPLOAD_BYTES = 200 * 1024 * 1024
MAX_TASK_PAGES = 200
TRIAL_PAGE_LIMIT = 10
FIELD_CONF_THRESHOLD = 0.70
FIELD_MIN_VALUE_CONFIDENCE = 0.50
LOCAL_QWEN_URL = os.environ.get("DOC_WORKBENCH_LOCAL_LLM_URL", "http://127.0.0.1:11435/v1/chat/completions")
DEFAULT_LOCAL_QWEN_MODEL = "Qwen2.5-0.5B-Instruct-GGUF-Q5_K_M"
DEFAULT_LOCAL_QWEN_MODEL_PATH = ROOT / "models" / "qwen2.5-0.5b-instruct-q5_k_m.gguf"
LOCAL_QWEN_MODEL = os.environ.get("DOC_WORKBENCH_LOCAL_LLM_DISPLAY_NAME", DEFAULT_LOCAL_QWEN_MODEL)
LOCAL_QWEN_MODEL_PATH = Path(os.environ.get("DOC_WORKBENCH_LOCAL_LLM_MODEL_PATH", str(DEFAULT_LOCAL_QWEN_MODEL_PATH)))
LOCAL_QWEN_REQUEST_MODEL = os.environ.get("DOC_WORKBENCH_LOCAL_LLM_MODEL", str(LOCAL_QWEN_MODEL_PATH))
LLAMA_SERVER_EXE = ROOT / "tools" / "llama" / ("llama-server.exe" if os.name == "nt" else "llama-server")
DASHSCOPE_CHAT_URL = os.environ.get("DASHSCOPE_CHAT_URL", "https://dashscope.aliyuncs.com/compatible-mode/v1/chat/completions")
PADDLE_VL_JOB_URL = os.environ.get("PADDLEOCR_VL_JOB_URL", "https://paddleocr.aistudio-app.com/api/v2/ocr/jobs")
PADDLE_VL_MODEL = "PaddleOCR-VL-1.6"

# ── 默认配置 ─────────────────────────────────────────────────────────────────
DEFAULT_CFG = {
    "model_variant": "tiny",     # tiny | small | medium
    "det_thresh": 0.20,
    "box_thresh": 0.40,
    "unclip":     1.40,
    "max_edge":   960,
    "min_size":    3,
    "provider":   "auto",        # auto | coreml | cpu
    "save_history": True,
    "thread_count": 4,
}

def load_config():
    if CFG_PATH.exists():
        try:
            return {**DEFAULT_CFG, **json.loads(CFG_PATH.read_text())}
        except Exception:
            pass
    return dict(DEFAULT_CFG)

def save_config(cfg: dict):
    merged = {**DEFAULT_CFG, **cfg}
    CFG_PATH.write_text(json.dumps(merged, indent=2))
    return merged

# ── 模型参数（固定） ────────────────────────────────────────────────────────
DET_MEAN  = np.array([0.485, 0.456, 0.406], dtype=np.float32)
DET_STD   = np.array([0.229, 0.224, 0.225], dtype=np.float32)
REC_H     = 48
REC_MAX_W = 2400

# ── 模型单例 ─────────────────────────────────────────────────────────────────
_det_sess = _rec_sess = _char_list = _current_provider = _current_model_variant = None
_qwen_sidecar_process = None


def available_model_variants():
    variants = {}
    for key, meta in MODEL_VARIANTS.items():
        root = meta["path"]
        ok = (root / "det" / "inference.onnx").exists() and (root / "rec" / "inference.onnx").exists()
        variants[key] = {
            "key": key,
            "label": meta["label"],
            "official_size": meta["official_size"],
            "use_case": meta["use_case"],
            "installed": ok,
        }
    return variants


def get_model_root(model_variant: str | None = None):
    key = model_variant or load_config().get("model_variant", "tiny")
    if key not in MODEL_VARIANTS:
        key = "tiny"
    root = MODEL_VARIANTS[key]["path"]
    if not ((root / "det" / "inference.onnx").exists() and (root / "rec" / "inference.onnx").exists()):
        key = "tiny"
        root = MODEL_VARIANTS[key]["path"]
    return key, root

def get_models(force_provider: str = None, force_model: str = None):
    global _det_sess, _rec_sess, _char_list, _current_provider, _current_model_variant
    cfg = load_config()
    want = force_provider or cfg.get("provider", "auto")
    model_key, model_root = get_model_root(force_model or cfg.get("model_variant", "tiny"))

    if _det_sess is not None and _current_provider == want and _current_model_variant == model_key:
        return _det_sess, _rec_sess, _char_list

    avail = ort.get_available_providers()
    prefer = []
    if want == "coreml" and "CoreMLExecutionProvider" in avail:
        prefer = ["CoreMLExecutionProvider", "CPUExecutionProvider"]
    elif want == "cpu":
        prefer = ["CPUExecutionProvider"]
    else:  # auto
        if "CoreMLExecutionProvider" in avail:
            prefer.append("CoreMLExecutionProvider")
        prefer.append("CPUExecutionProvider")

    opts = ort.SessionOptions()
    opts.inter_op_num_threads = int(cfg.get("thread_count", 4))
    opts.intra_op_num_threads = int(cfg.get("thread_count", 4))

    _det_sess = ort.InferenceSession(str(model_root/"det"/"inference.onnx"),
                                     sess_options=opts, providers=prefer)
    _rec_sess = ort.InferenceSession(str(model_root/"rec"/"inference.onnx"),
                                     sess_options=opts, providers=prefer)
    char_json = model_root / "rec" / "char_dict.json"
    if char_json.exists():
        with open(char_json, encoding="utf-8") as f:
            d = json.load(f)
    else:
        rec_cfg = yaml.safe_load((model_root / "rec" / "inference.yml").read_text(encoding="utf-8"))
        d = rec_cfg.get("PostProcess", {}).get("character_dict", [])
        if not d:
            raise RuntimeError(f"Cannot load OCR character dictionary from {model_root / 'rec'}")
    _char_list = [""] + d + [" "]
    _current_provider = want
    _current_model_variant = model_key
    return _det_sess, _rec_sess, _char_list

def current_backend():
    if _det_sess is None:
        return "未加载"
    return _det_sess.get_providers()[0]

def current_model_variant():
    return _current_model_variant or load_config().get("model_variant", "tiny")

def content_disposition(filename: str) -> str:
    """RFC 5987 filename header with an ASCII fallback for Chinese filenames."""
    clean = (filename or "ocr").replace("\n", "_").replace("\r", "_").replace('"', "_")
    fallback = "".join(ch if 32 <= ord(ch) < 127 else "_" for ch in clean) or "ocr"
    return f'attachment; filename="{fallback}"; filename*=UTF-8\'\'{quote(clean)}'

# ── 数据库 ───────────────────────────────────────────────────────────────────
def db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    with db() as c:
        c.execute("""CREATE TABLE IF NOT EXISTS history(
            id          TEXT PRIMARY KEY,
            created_at  REAL NOT NULL,
            filename    TEXT NOT NULL,
            n_boxes     INT,
            det_ms      INT,
            rec_ms      INT,
            total_ms    INT,
            backend     TEXT,
            text        TEXT,
            lines_json  TEXT,
            tables_json TEXT,
            upload_path TEXT,
            thumb_path  TEXT,
            annotated_path TEXT
        )""")
        # Migrate older schema (add tables_json if missing)
        cols = [r[1] for r in c.execute("PRAGMA table_info(history)").fetchall()]
        if "tables_json" not in cols:
            c.execute("ALTER TABLE history ADD COLUMN tables_json TEXT")
        c.execute("""CREATE TABLE IF NOT EXISTS tasks(
            id              TEXT PRIMARY KEY,
            created_at      REAL NOT NULL,
            updated_at      REAL NOT NULL,
            title           TEXT NOT NULL,
            mode            TEXT NOT NULL,
            status          TEXT NOT NULL,
            total_pages     INT NOT NULL DEFAULT 0,
            processed_pages INT NOT NULL DEFAULT 0,
            failed_pages    INT NOT NULL DEFAULT 0,
            source_names_json TEXT,
            fields_json     TEXT,
            has_fields      INT NOT NULL DEFAULT 0,
            source_bytes    INT NOT NULL DEFAULT 0,
            trial_pages_used INT NOT NULL DEFAULT 0,
            error           TEXT
        )""")
        c.execute("""CREATE TABLE IF NOT EXISTS pages(
            id              TEXT PRIMARY KEY,
            task_id         TEXT NOT NULL,
            page_index      INT NOT NULL,
            source_name     TEXT NOT NULL,
            page_number     INT NOT NULL,
            status          TEXT NOT NULL,
            text            TEXT,
            lines_json      TEXT,
            backend         TEXT,
            det_ms          INT,
            rec_ms          INT,
            total_ms        INT,
            image_path      TEXT,
            thumb_path      TEXT,
            annotated_path  TEXT,
            extractor       TEXT,
            error           TEXT,
            FOREIGN KEY(task_id) REFERENCES tasks(id)
        )""")
        page_cols = [r[1] for r in c.execute("PRAGMA table_info(pages)").fetchall()]
        if "html" not in page_cols:
            c.execute("ALTER TABLE pages ADD COLUMN html TEXT")
        if "markdown" not in page_cols:
            c.execute("ALTER TABLE pages ADD COLUMN markdown TEXT")
        c.execute("""CREATE TABLE IF NOT EXISTS field_results(
            id          TEXT PRIMARY KEY,
            task_id     TEXT NOT NULL,
            page_id     TEXT NOT NULL,
            field_name  TEXT NOT NULL,
            value       TEXT,
            confidence  REAL,
            evidence    TEXT,
            status      TEXT,
            FOREIGN KEY(task_id) REFERENCES tasks(id),
            FOREIGN KEY(page_id) REFERENCES pages(id)
        )""")
        c.execute("""CREATE TABLE IF NOT EXISTS usage(
            key     TEXT PRIMARY KEY,
            value   INT NOT NULL
        )""")
        c.commit()

# ── 推理逻辑 ─────────────────────────────────────────────────────────────────
def det_preprocess(img, max_edge):
    w, h = img.size
    r = min(1.0, max_edge / max(w, h))
    nw = max(32, round(w * r / 32) * 32)
    nh = max(32, round(h * r / 32) * 32)
    arr = np.array(img.resize((nw, nh), Image.LANCZOS).convert("RGB"),
                   dtype=np.float32) / 255.0
    arr = (arr - DET_MEAN) / DET_STD
    return arr.transpose(2, 0, 1)[np.newaxis], w/nw, h/nh, nw, nh

def flood_boxes(prob, pw, ph, sx, sy, det_thresh, box_thresh, unclip, min_size):
    binary   = (prob > det_thresh).astype(np.uint8).ravel()
    labeled  = np.zeros(pw * ph, dtype=np.int32)
    cur = 0
    boxes = []
    for start in range(pw * ph):
        if binary[start] != 1 or labeled[start] != 0:
            continue
        cur += 1
        stack = [start]
        labeled[start] = cur
        xs, ys, vals = [], [], []
        while stack:
            p = stack.pop()
            x, y = p % pw, p // pw
            xs.append(x); ys.append(y); vals.append(prob[y, x])
            for nx2, ny2 in ((x-1,y),(x+1,y),(x,y-1),(x,y+1)):
                if 0 <= nx2 < pw and 0 <= ny2 < ph:
                    idx = ny2 * pw + nx2
                    if binary[idx] == 1 and labeled[idx] == 0:
                        labeled[idx] = cur
                        stack.append(idx)
        bw = max(xs)-min(xs)+1; bh = max(ys)-min(ys)+1
        if min(bw, bh) < min_size: continue
        if float(np.mean(vals)) < box_thresh: continue
        d = bw * bh * unclip / (2 * (bw + bh))
        boxes.append(dict(
            x0=max(0,(min(xs)-d)*sx), y0=max(0,(min(ys)-d)*sy),
            x1=(max(xs)+d)*sx,        y1=(max(ys)+d)*sy,
            cy=(min(ys)+max(ys))/2*sy
        ))
    boxes.sort(key=lambda b: (round(b["cy"]/10)*10, b["x0"]))
    return boxes


def order_points_clockwise(pts):
    pts = np.asarray(pts, dtype=np.float32)
    if pts.shape != (4, 2):
        return pts

    x_span = float(pts[:, 0].max() - pts[:, 0].min())
    y_span = float(pts[:, 1].max() - pts[:, 1].min())
    ordered = np.zeros((4, 2), dtype=np.float32)

    # For long OCR text boxes, the classic sum/diff ordering can collapse when
    # a slanted box touches an image edge. Split the box into left/right (or
    # top/bottom for vertical text) pairs instead; this preserves all corners.
    if x_span >= y_span:
        by_x = pts[np.argsort(pts[:, 0])]
        left = by_x[:2][np.argsort(by_x[:2, 1])]
        right = by_x[2:][np.argsort(by_x[2:, 1])]
        ordered[0], ordered[3] = left[0], left[1]
        ordered[1], ordered[2] = right[0], right[1]
    else:
        by_y = pts[np.argsort(pts[:, 1])]
        top = by_y[:2][np.argsort(by_y[:2, 0])]
        bottom = by_y[2:][np.argsort(by_y[2:, 0])]
        ordered[0], ordered[1] = top[0], top[1]
        ordered[3], ordered[2] = bottom[0], bottom[1]

    return ordered


def contour_boxes(prob, pw, ph, sx, sy, det_thresh, box_thresh, unclip, min_size):
    """DB-style rotated boxes from the probability map.

    The first local version used axis-aligned connected components. That is fast,
    but it breaks badly on perspective/rotated text such as business cards. This
    contour path keeps the text angle and lets recognition crop with perspective
    rectification, closer to PaddleOCR's official pipeline.
    """
    if cv2 is None:
        return flood_boxes(prob, pw, ph, sx, sy, det_thresh, box_thresh, unclip, min_size)

    bitmap = (prob > det_thresh).astype(np.uint8)
    kernel = np.ones((2, 2), dtype=np.uint8)
    bitmap = cv2.dilate(bitmap, kernel, iterations=1)
    contours, _ = cv2.findContours(bitmap * 255, cv2.RETR_LIST, cv2.CHAIN_APPROX_SIMPLE)

    boxes = []
    for contour in contours:
        if contour.shape[0] < 3:
            continue
        rect = cv2.minAreaRect(contour)
        (cx, cy), (w, h), _ = rect
        if min(w, h) < min_size:
            continue

        pts = cv2.boxPoints(rect).astype(np.float32)
        mask = np.zeros(prob.shape, dtype=np.uint8)
        cv2.fillPoly(mask, [pts.astype(np.int32)], 1)
        score = float(cv2.mean(prob, mask=mask)[0])
        if score < box_thresh:
            continue

        raw_poly = np.column_stack([pts[:, 0] * sx, pts[:, 1] * sy]).astype(float)
        ordered = order_points_clockwise(raw_poly)
        edge = ordered[1] - ordered[0]
        angle = math.degrees(math.atan2(float(edge[1]), float(edge[0])))
        while angle <= -90:
            angle += 180
        while angle > 90:
            angle -= 180

        center = pts.mean(axis=0, keepdims=True)
        # Approximate DB unclip for a rectangle. Horizontal text usually needs
        # much less padding than slanted perspective text; over-expanding it can
        # pull in background texture and hurt recognition.
        base_scale = max(1.0, min(float(unclip), 2.2))
        scale = min(base_scale, 1.08) if abs(angle) < 10 else base_scale
        pts = center + (pts - center) * scale
        pts[:, 0] = np.clip(pts[:, 0], 0, pw - 1)
        pts[:, 1] = np.clip(pts[:, 1], 0, ph - 1)

        poly = np.column_stack([pts[:, 0] * sx, pts[:, 1] * sy]).astype(float)
        x0, y0 = poly.min(axis=0)
        x1, y1 = poly.max(axis=0)
        if min(x1 - x0, y1 - y0) < 2:
            continue
        boxes.append({
            "x0": float(x0), "y0": float(y0),
            "x1": float(x1), "y1": float(y1),
            "cy": float(poly[:, 1].mean()),
            "poly": poly.tolist(),
            "angle": float(angle),
            "score": score,
        })

    boxes.sort(key=lambda b: (round(b["cy"] / 10) * 10, b["x0"]))
    return boxes


def rec_tensor_from_crop(crop: Image.Image):
    if crop.width <= 1 or crop.height <= 1:
        return None
    nw = max(8, min(REC_MAX_W, round(crop.width * REC_H / crop.height)))
    arr = np.array(crop.resize((nw, REC_H), Image.LANCZOS), dtype=np.float32)/255.0
    return ((arr - 0.5) / 0.5).transpose(2,0,1)[np.newaxis]


def affine_rec_crop(img, b, pts):
    arr = np.array(img.convert("RGB"))
    h, w = arr.shape[:2]
    angle = float(b.get("angle", 0))
    mat = cv2.getRotationMatrix2D((w / 2, h / 2), angle, 1.0)
    cos = abs(mat[0, 0])
    sin = abs(mat[0, 1])
    dst_w = int((h * sin) + (w * cos))
    dst_h = int((h * cos) + (w * sin))
    mat[0, 2] += dst_w / 2 - w / 2
    mat[1, 2] += dst_h / 2 - h / 2
    rotated = cv2.warpAffine(arr, mat, (dst_w, dst_h), borderMode=cv2.BORDER_REPLICATE)
    rpts = np.hstack([pts, np.ones((4, 1), dtype=np.float32)]) @ mat.T
    x0, y0 = rpts.min(axis=0)
    x1, y1 = rpts.max(axis=0)
    bw = max(1.0, float(x1 - x0))
    bh = max(1.0, float(y1 - y0))
    x0 = max(0, int(x0 - bw * 0.02))
    x1 = min(dst_w, int(x1 + bw * 0.02))
    y0 = max(0, int(y0 - bh * 0.30))
    y1 = min(dst_h, int(y1 + bh * 0.30))
    if x1 <= x0 or y1 <= y0:
        return None
    crop = Image.fromarray(rotated[y0:y1, x0:x1]).convert("RGB")
    return rec_tensor_from_crop(crop)


def rec_crop(img, b):
    # Near-horizontal text is usually cleaner with a normal crop. Perspective
    # warping helps slanted business cards, but can distort large dot-matrix
    # letters that were already horizontal.
    if b.get("poly") is not None and cv2 is not None and abs(float(b.get("angle", 0))) >= 10:
        pts = order_points_clockwise(np.array(b["poly"], dtype=np.float32))
        width_a = np.linalg.norm(pts[2] - pts[3])
        width_b = np.linalg.norm(pts[1] - pts[0])
        height_a = np.linalg.norm(pts[1] - pts[2])
        height_b = np.linalg.norm(pts[0] - pts[3])
        long_side = max(width_a, width_b)
        short_side = max(1.0, max(height_a, height_b))
        if long_side / short_side >= 6:
            rotated_crop = affine_rec_crop(img, b, pts)
            if rotated_crop is not None:
                return rotated_crop
        dst_w = max(8, int(round(max(width_a, width_b))))
        dst_h = max(8, int(round(max(height_a, height_b))))
        dst = np.array([[0, 0], [dst_w - 1, 0], [dst_w - 1, dst_h - 1], [0, dst_h - 1]],
                       dtype=np.float32)
        arr = np.array(img.convert("RGB"))
        mat = cv2.getPerspectiveTransform(pts, dst)
        warped = cv2.warpPerspective(arr, mat, (dst_w, dst_h), borderMode=cv2.BORDER_REPLICATE)
        if warped.shape[0] / max(1, warped.shape[1]) >= 1.5:
            warped = np.rot90(warped)
        crop = Image.fromarray(warped).convert("RGB")
        return rec_tensor_from_crop(crop)

    x0,y0,x1,y1 = (max(0,int(b[k])) for k in ("x0","y0","x1","y1"))
    x1 = min(x1, img.width); y1 = min(y1, img.height)
    if x1<=x0 or y1<=y0: return None
    crop = img.crop((x0,y0,x1,y1)).convert("RGB")
    return rec_tensor_from_crop(crop)

def ctc_decode(logits, char_list):
    prev, out = -1, ""
    for idx in logits.argmax(axis=1):
        if idx != 0 and idx != prev:
            out += char_list[idx] if idx < len(char_list) else ""
        prev = idx
    return out

def run_ocr(img: Image.Image, cfg: dict):
    det, rec, chars = get_models(force_model=cfg.get("model_variant"))
    t0 = time.perf_counter()
    tensor, sx, sy, pw, ph = det_preprocess(img, cfg["max_edge"])
    prob = det.run(None, {det.get_inputs()[0].name: tensor})[0][0, 0]
    det_ms = (time.perf_counter() - t0) * 1000
    boxes = contour_boxes(prob, pw, ph, sx, sy,
                          cfg["det_thresh"], cfg["box_thresh"],
                          cfg["unclip"], cfg["min_size"])
    t1 = time.perf_counter()
    lines = []
    for b in boxes:
        crop = rec_crop(img, b)
        if crop is None: continue
        out = rec.run(None, {rec.get_inputs()[0].name: crop})[0][0]
        txt = ctc_decode(out, chars)
        if txt.strip():
            item = {"text": txt, "box": [b["x0"],b["y0"],b["x1"],b["y1"]]}
            if b.get("poly") is not None:
                item["poly"] = b["poly"]
            lines.append(item)
    rec_ms = (time.perf_counter() - t1) * 1000
    return lines, det_ms, rec_ms

# ── 表格识别（启发式：行Y聚类 + 列X聚类） ──────────────────────────────────
def _median(vals, default=0):
    vals = sorted(v for v in vals if v is not None)
    if not vals:
        return default
    return vals[len(vals) // 2]


def _cluster_positions(values, tol):
    groups = []
    for v in sorted(values):
        if not groups or abs(v - groups[-1][-1]) > tol:
            groups.append([v])
        else:
            groups[-1].append(v)
    return [sum(g) / len(g) for g in groups]


def detect_tables(lines: list):
    """从 OCR 文本框还原表格结构。

    这不是 PP-Structure/SLANet 那类深度表格结构模型，而是面向本地 Tiny
    OCR 的几何重建：先聚成文本行，再用 X 方向中心点聚成列。相比上一版仅
    依赖每行文字块数量的算法，这版对中文表格、缺失单元格和列数轻微变化
    更宽容。
    返回: [{rows, n_rows, n_cols, bbox, confidence}]
    """
    if len(lines) < 4:
        return []

    items = [{
        "text": str(l.get("text", "")).strip(),
        "x0": l["box"][0], "y0": l["box"][1],
        "x1": l["box"][2], "y1": l["box"][3],
        "cx": (l["box"][0] + l["box"][2]) / 2,
        "cy": (l["box"][1] + l["box"][3]) / 2,
        "w":  max(1, l["box"][2] - l["box"][0]),
        "h":  max(1, l["box"][3] - l["box"][1]),
    } for l in lines if l.get("text", "").strip() and len(l.get("box", [])) == 4]
    if len(items) < 4:
        return []

    items.sort(key=lambda x: x["cy"])
    median_h = _median([it["h"] for it in items], 12)

    # 行聚类
    rows = [[items[0]]]
    for it in items[1:]:
        row_cy = sum(x["cy"] for x in rows[-1]) / len(rows[-1])
        avg_h = (it["h"] + _median([x["h"] for x in rows[-1]], median_h)) / 2
        if abs(it["cy"] - row_cy) <= max(8, avg_h * 0.8):
            rows[-1].append(it)
        else:
            rows.append([it])
    for r in rows:
        r.sort(key=lambda x: x["x0"])

    # 找候选表格区段
    tables = []
    i = 0
    while i < len(rows):
        if len(rows[i]) < 2:
            i += 1; continue
        j = i + 1
        while j < len(rows) and len(rows[j]) >= 2:
            prev_y = sum(x["cy"] for x in rows[j-1]) / len(rows[j-1])
            cur_y = sum(x["cy"] for x in rows[j]) / len(rows[j])
            if cur_y - prev_y > median_h * 4.0:
                break
            j += 1

        run = rows[i:j]
        if len(run) >= 2:
            median_w = _median([c["w"] for r in run for c in r], 24)
            xs = [c["cx"] for r in run for c in r]
            col_centers = _cluster_positions(xs, max(18, median_w * 0.65))
            col_centers = sorted(col_centers)

            # 过密的列通常是同一单元格被切成多个短词，二次合并一下。
            if len(col_centers) > 2:
                gaps = [b - a for a, b in zip(col_centers, col_centers[1:])]
                med_gap = _median(gaps, 0)
                if med_gap:
                    col_centers = _cluster_positions(col_centers, max(18, med_gap * 0.28))

            max_cols = len(col_centers)
            aligned_rows = 0
            table_rows = []
            for r in run:
                cells = [""] * max_cols
                used = set()
                for cell in r:
                    ci = min(range(max_cols),
                             key=lambda k: abs(cell["cx"] - col_centers[k]))
                    used.add(ci)
                    cells[ci] = (cells[ci] + " " + cell["text"]).strip() \
                                if cells[ci] else cell["text"]
                if len(used) >= min(2, max_cols):
                    aligned_rows += 1
                table_rows.append(cells)

            enough_rows = len(run) >= 3 or (len(run) >= 2 and max_cols >= 3)
            confidence = aligned_rows / max(1, len(run))
            if max_cols >= 2 and enough_rows and confidence >= 0.75:
                xs2 = [c["x0"] for r in run for c in r] + [c["x1"] for r in run for c in r]
                ys2 = [c["y0"] for r in run for c in r] + [c["y1"] for r in run for c in r]
                tables.append({
                    "rows":   table_rows,
                    "n_rows": len(table_rows),
                    "n_cols": max_cols,
                    "bbox":   [min(xs2), min(ys2), max(xs2), max(ys2)],
                    "confidence": round(confidence, 3),
                })
        i = max(j, i + 1)
    return tables


def tables_to_csv(tables: list) -> str:
    buf = io.StringIO()
    buf.write("\ufeff")
    for idx, t in enumerate(tables):
        if idx > 0: buf.write("\n\n")
        w = csv.writer(buf)
        for row in t.get("rows", []):
            w.writerow(row)
    return buf.getvalue()


def tables_to_markdown(tables: list) -> str:
    parts = []
    for idx, t in enumerate(tables):
        rows = t["rows"]
        if not rows: continue
        parts.append(f"### 表格 {idx+1}\n")
        # 第一行作表头
        header = rows[0]
        parts.append("| " + " | ".join(c.replace("|","\\|") for c in header) + " |")
        parts.append("| " + " | ".join(["---"] * len(header)) + " |")
        for row in rows[1:]:
            row = row + [""] * (len(header) - len(row))  # pad
            parts.append("| " + " | ".join(c.replace("|","\\|") for c in row[:len(header)]) + " |")
        parts.append("")
    return "\n".join(parts)


def tables_to_html(tables: list) -> str:
    parts = ['<!DOCTYPE html><html><head><meta charset="utf-8">',
             '<style>body{font-family:-apple-system,sans-serif;padding:24px;max-width:960px;margin:0 auto}',
             'h2{font-size:16px;margin:24px 0 10px;color:#333}',
             'table{border-collapse:collapse;width:100%;margin-bottom:18px;font-size:13px}',
             'th,td{border:1px solid #ddd;padding:7px 11px;text-align:left}',
             'th{background:#f5f5f7;font-weight:600}',
             'tr:nth-child(even) td{background:#fafafa}</style></head><body>']
    for idx, t in enumerate(tables):
        parts.append(f"<h2>表格 {idx+1} · {t.get('n_rows', 0)}×{t.get('n_cols', 0)}</h2>")
        parts.append("<table>")
        for ri, row in enumerate(t.get("rows", [])):
            tag = "th" if ri == 0 else "td"
            parts.append("<tr>" + "".join(
                f"<{tag}>{html.escape(str(c))}</{tag}>"
                for c in row) + "</tr>")
        parts.append("</table>")
    parts.append("</body></html>")
    return "\n".join(parts)


def tables_to_xlsx(tables: list) -> bytes:
    if Workbook is None:
        raise HTTPException(500, "openpyxl is not installed; cannot export xlsx")
    wb = Workbook()
    wb.remove(wb.active)
    header_fill = PatternFill("solid", fgColor="F2F4F7")
    header_font = Font(bold=True, color="111827")
    thin = Side(style="thin", color="D0D5DD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    if not tables:
        ws = wb.create_sheet("OCR")
        ws["A1"] = ""
    for idx, t in enumerate(tables):
        ws = wb.create_sheet(f"表格{idx+1}")
        rows = t.get("rows", [])
        for r_idx, row in enumerate(rows, start=1):
            for c_idx, value in enumerate(row, start=1):
                cell = ws.cell(r_idx, c_idx, value=value)
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                if r_idx == 1:
                    cell.fill = header_fill
                    cell.font = header_font
        for c_idx in range(1, (t.get("n_cols") or (len(rows[0]) if rows else 1)) + 1):
            letter = get_column_letter(c_idx)
            max_len = max([len(str(ws.cell(r, c_idx).value or "")) for r in range(1, ws.max_row + 1)] or [8])
            ws.column_dimensions[letter].width = min(max(10, max_len * 1.7), 42)
        ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def annotate_image(img: Image.Image, lines: list, save_path: Path = None):
    vis = img.convert("RGB").copy()
    draw = ImageDraw.Draw(vis)
    for item in lines:
        if item.get("poly"):
            pts = [(int(x), int(y)) for x, y in item["poly"]]
            draw.line(pts + [pts[0]], fill=(124,92,255), width=2)
        else:
            x0,y0,x1,y1 = (int(v) for v in item["box"])
            draw.rectangle([x0,y0,x1,y1], outline=(124,92,255), width=2)
    vis.thumbnail((1600, 1600), Image.LANCZOS)
    if save_path:
        vis.save(save_path, "JPEG", quality=85)
    buf = io.BytesIO()
    vis.save(buf, "JPEG", quality=82)
    return "data:image/jpeg;base64," + base64.b64encode(buf.getvalue()).decode()

def make_thumb(img: Image.Image, save_path: Path):
    thumb = img.convert("RGB").copy()
    thumb.thumbnail((160, 160), Image.LANCZOS)
    thumb.save(save_path, "JPEG", quality=78)


# Document Workbench helpers
def _now():
    return time.time()


def safe_name(name: str) -> str:
    name = (name or "document").replace("\\", "/").split("/")[-1]
    return re.sub(r"[^\w\u4e00-\u9fff.\-()（）]+", "_", name).strip("_") or "document"


def public_path(path: Path | str | None) -> str:
    if not path:
        return ""
    p = Path(path)
    try:
        return str(p.relative_to(ROOT / "webapp")).replace("\\", "/")
    except Exception:
        return str(p).replace("\\", "/")


def parse_json_object(text: str):
    text = (text or "").strip()
    try:
        return json.loads(text)
    except Exception:
        start = text.find("{")
        end = text.rfind("}")
        if start >= 0 and end > start:
            snippet = text[start:end + 1]
            try:
                return json.loads(snippet)
            except Exception:
                pass
        if start >= 0:
            snippet = text[start:].strip()
            # Small local models sometimes stop one or two closing braces early.
            # Only repair structural tails; never invent content.
            stack = []
            in_str = False
            esc_next = False
            for ch in snippet:
                if esc_next:
                    esc_next = False
                    continue
                if ch == "\\" and in_str:
                    esc_next = True
                    continue
                if ch == '"':
                    in_str = not in_str
                    continue
                if in_str:
                    continue
                if ch in "{[":
                    stack.append(ch)
                elif ch in "}]":
                    if stack and ((stack[-1] == "{" and ch == "}") or (stack[-1] == "[" and ch == "]")):
                        stack.pop()
            if not in_str and stack:
                tail = "".join("}" if ch == "{" else "]" for ch in reversed(stack))
                return json.loads(snippet + tail)
        raise


def normalize_fields(raw) -> list[dict]:
    if raw is None:
        return []
    if isinstance(raw, str):
        try:
            raw = json.loads(raw)
        except Exception:
            raw = [x.strip() for x in raw.split(",") if x.strip()]
    fields = []
    seen = set()
    for item in raw if isinstance(raw, list) else []:
        if isinstance(item, str):
            name = item.strip()
        elif isinstance(item, dict):
            name = str(item.get("name") or item.get("field") or item.get("title") or "").strip()
        else:
            name = ""
        if name and name not in seen:
            fields.append({"name": name})
            seen.add(name)
    return fields


def fields_from_excel(data: bytes) -> list[dict]:
    if load_workbook is None:
        raise HTTPException(500, "openpyxl is not installed; cannot read Excel templates")
    try:
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        ws = wb.worksheets[0]
        first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    except Exception:
        raise HTTPException(400, "无法读取 Excel 模板，请确认文件格式正确")
    names = [str(v).strip() for v in (first_row or []) if v is not None and str(v).strip()]
    if not names:
        raise HTTPException(400, "Excel 模板第一行没有可用字段名")
    return normalize_fields(names)


class _PlainTextHTMLParser(HTMLParser):
    def __init__(self):
        super().__init__(convert_charrefs=True)
        self.parts = []

    def handle_starttag(self, tag, attrs):
        if tag in ("tr", "p", "div", "section", "article", "h1", "h2", "h3", "h4", "h5", "h6", "li"):
            self.parts.append("\n")
        elif tag in ("td", "th"):
            self.parts.append("\t")
        elif tag == "br":
            self.parts.append("\n")

    def handle_endtag(self, tag):
        if tag in ("tr", "p", "div", "li", "table"):
            self.parts.append("\n")

    def handle_data(self, data):
        if data:
            self.parts.append(data)

    def text(self):
        raw = "".join(self.parts)
        lines = []
        for line in raw.splitlines():
            cells = [c.strip() for c in line.split("\t")]
            compact = "\t".join(c for c in cells if c)
            if compact:
                lines.append(compact)
        return "\n".join(lines).strip()


class _TableCellHTMLParser(HTMLParser):
    def __init__(self):
        super().__init__(convert_charrefs=True)
        self.rows = []
        self.current_row = None
        self.current_cell = None
        self.in_cell = False

    def handle_starttag(self, tag, attrs):
        tag = tag.lower()
        if tag == "tr":
            self.current_row = []
        elif tag in ("td", "th") and self.current_row is not None:
            self.current_cell = []
            self.in_cell = True
        elif tag == "br" and self.in_cell:
            self.current_cell.append("\n")

    def handle_endtag(self, tag):
        tag = tag.lower()
        if tag in ("td", "th") and self.current_row is not None and self.current_cell is not None:
            self.current_row.append("".join(self.current_cell).strip())
            self.current_cell = None
            self.in_cell = False
        elif tag == "tr" and self.current_row is not None:
            self.rows.append(self.current_row)
            self.current_row = None

    def handle_data(self, data):
        if self.in_cell and self.current_cell is not None:
            self.current_cell.append(data)


def html_table_key_value_text(html_text: str) -> str:
    parser = _TableCellHTMLParser()
    try:
        parser.feed(html_text or "")
    except Exception:
        return ""
    lines = []
    for row in parser.rows:
        for i in range(0, max(0, len(row) - 1), 2):
            key = (row[i] or "").strip()
            val = (row[i + 1] or "").strip()
            if key:
                lines.append(f"{key}：{val}")
    return "\n".join(lines)


class _SanitizingHTMLParser(HTMLParser):
    ALLOWED_TAGS = {
        "table", "thead", "tbody", "tfoot", "tr", "td", "th", "colgroup", "col",
        "p", "br", "strong", "b", "em", "i", "u", "span", "div", "ul", "ol", "li",
        "h1", "h2", "h3", "h4", "h5", "h6",
    }
    ALLOWED_ATTRS = {"colspan", "rowspan", "scope", "align", "border"}

    def __init__(self):
        super().__init__(convert_charrefs=True)
        self.parts = []

    def handle_starttag(self, tag, attrs):
        tag = tag.lower()
        if tag not in self.ALLOWED_TAGS:
            return
        clean_attrs = []
        for key, value in attrs:
            key = (key or "").lower()
            if key in self.ALLOWED_ATTRS and value is not None:
                clean_attrs.append(f' {key}="{html.escape(str(value), quote=True)}"')
        self.parts.append(f"<{tag}{''.join(clean_attrs)}>")

    def handle_endtag(self, tag):
        tag = tag.lower()
        if tag in self.ALLOWED_TAGS:
            self.parts.append(f"</{tag}>")

    def handle_startendtag(self, tag, attrs):
        tag = tag.lower()
        if tag == "br":
            self.parts.append("<br>")

    def handle_data(self, data):
        self.parts.append(html.escape(data))

    def html(self):
        return "".join(self.parts).strip()


def html_to_plain_text(html_text: str) -> str:
    parser = _PlainTextHTMLParser()
    try:
        parser.feed(html_text or "")
        return parser.text()
    except Exception:
        return re.sub(r"<[^>]+>", " ", html_text or "").strip()


def sanitize_result_html(html_text: str) -> str:
    if not html_text or "<" not in html_text:
        return ""
    parser = _SanitizingHTMLParser()
    try:
        parser.feed(html_text)
        return parser.html()
    except Exception:
        return ""


def download_cloud_image(url: str, path: Path) -> bool:
    if not url or requests is None:
        return False
    try:
        r = requests.get(url, timeout=60)
        r.raise_for_status()
        img = Image.open(io.BytesIO(r.content)).convert("RGB")
        img.thumbnail((1600, 1600), Image.LANCZOS)
        img.save(path, "JPEG", quality=88)
        return True
    except Exception:
        return False


def load_secret_blob() -> dict:
    if not SECRET_BLOB_PATH.exists():
        return {}
    try:
        raw = base64.b64decode(SECRET_BLOB_PATH.read_text(encoding="utf-8").strip())
        key = b"doc-workbench-v1"
        decoded = bytes(b ^ key[i % len(key)] for i, b in enumerate(raw))
        return json.loads(decoded.decode("utf-8"))
    except Exception:
        return {}


def get_secret(name: str) -> str:
    val = os.environ.get(name)
    if val:
        return val
    return str(load_secret_blob().get(name) or "")


def get_usage_pages() -> int:
    with db() as c:
        row = c.execute("SELECT value FROM usage WHERE key='trial_pages_used'").fetchone()
        return int(row["value"]) if row else 0


def set_usage_pages(count: int):
    with db() as c:
        c.execute("""INSERT INTO usage(key,value) VALUES('trial_pages_used',?)
                     ON CONFLICT(key) DO UPDATE SET value=excluded.value""",
                  (int(count),))
        c.commit()


def ensure_trial_available(page_count: int):
    used = get_usage_pages()
    if used + page_count > TRIAL_PAGE_LIMIT:
        raise HTTPException(
            402,
            f"试用额度不足：共 {TRIAL_PAGE_LIMIT} 页，已使用 {used} 页，本次需要 {page_count} 页"
        )


def image_from_bytes(data: bytes, filename: str) -> Image.Image:
    try:
        return Image.open(io.BytesIO(data)).convert("RGB")
    except Exception:
        raise HTTPException(400, f"无法解析图片：{filename}")


def render_pdf_pages(data: bytes, filename: str) -> list[Image.Image]:
    if pdfium is None:
        raise HTTPException(500, "PDF 支持需要安装 pypdfium2")
    tmp_name = None
    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as f:
            f.write(data)
            tmp_name = f.name
        pdf = pdfium.PdfDocument(tmp_name)
        images = []
        for i in range(len(pdf)):
            page = pdf[i]
            bitmap = page.render(scale=2.0)
            images.append(bitmap.to_pil().convert("RGB"))
        try:
            pdf.close()
        except Exception:
            pass
        if not images:
            raise HTTPException(400, f"PDF 没有可处理页面：{filename}")
        return images
    except HTTPException:
        raise
    except Exception:
        raise HTTPException(400, f"PDF 渲染失败：{filename}")
    finally:
        if tmp_name:
            try:
                os.unlink(tmp_name)
            except Exception:
                pass


def build_page_inputs(uploaded: list[tuple[str, bytes]]) -> list[dict]:
    total_bytes = sum(len(data) for _, data in uploaded)
    if total_bytes > MAX_UPLOAD_BYTES:
        raise HTTPException(413, "单次上传不能超过 200MB")
    pages = []
    for filename, data in uploaded:
        clean = safe_name(filename)
        ext = clean.rsplit(".", 1)[-1].lower() if "." in clean else ""
        if ext == "pdf":
            images = render_pdf_pages(data, clean)
            for idx, img in enumerate(images, start=1):
                pages.append({"source_name": clean, "page_number": idx, "image": img})
        else:
            pages.append({"source_name": clean, "page_number": 1, "image": image_from_bytes(data, clean)})
        if len(pages) > MAX_TASK_PAGES:
            raise HTTPException(413, "单次上传不能超过 200 页")
    if not pages:
        raise HTTPException(400, "请选择 PDF 或图片文件")
    return pages


def compact_ocr_lines(lines: list[dict] | None) -> list[dict]:
    items = []
    for idx, line in enumerate(lines or [], start=1):
        txt = str(line.get("text") or "").strip()
        box = line.get("box") or []
        if not txt or len(box) != 4:
            continue
        try:
            x0, y0, x1, y1 = [round(float(v), 1) for v in box]
        except Exception:
            continue
        items.append({
            "id": f"L{idx:03d}",
            "text": txt,
            "box": [x0, y0, x1, y1],
            "cx": round((x0 + x1) / 2, 1),
            "cy": round((y0 + y1) / 2, 1),
        })
    items.sort(key=lambda x: (round(x["cy"] / 10) * 10, x["box"][0]))
    return items


def same_row(a: dict, b: dict) -> bool:
    ah = max(1.0, a["box"][3] - a["box"][1])
    bh = max(1.0, b["box"][3] - b["box"][1])
    return abs(a["cy"] - b["cy"]) <= max(8.0, (ah + bh) * 0.45)


def field_candidates(text: str, fields: list[dict], lines: list[dict] | None = None) -> dict:
    ocr_lines = compact_ocr_lines(lines)
    field_names = {f["name"] for f in fields}
    candidates: dict[str, list[dict]] = {f["name"]: [] for f in fields}

    def adjust_candidate(name: str, value: str, reason: str, confidence: float):
        value = str(value or "").strip(" \t\r\n:：,，;；")
        for other in field_names:
            if other != name and other in value:
                value = value.split(other, 1)[0].strip(" \t\r\n:：,，;；")
        if not value:
            return "", 0.0

        if any(k in name for k in ("日期", "时间")):
            m = re.search(r"\d{4}\s*年\s*\d{1,2}\s*月\s*\d{1,2}\s*日|\d{4}[-/.]\d{1,2}[-/.]\d{1,2}", value)
            if m:
                return m.group(0).replace(" ", ""), max(confidence, 0.86)
            return value, min(confidence, 0.35)

        if any(k in name for k in ("班级", "班号")):
            m = re.search(r"[\u4e00-\u9fffA-Za-z]{0,8}\s*\d{2,4}\s*(?:班)?", value)
            if m and not re.search(r"\d{4}\s*年", value):
                return m.group(0).strip(), max(confidence, 0.82)
            return value, min(confidence, 0.45)

        if any(k in name for k in ("教材", "书名", "用书")):
            if "《" in value:
                value = value[value.find("《"):].strip()
                return value, max(confidence, 0.84)
            if "出版社" in value:
                return value, max(confidence, 0.78)
            return value, min(confidence, 0.45)

        if any(k in name for k in ("教师", "姓名", "负责人", "联系人")):
            if re.search(r"\d|授课|日期|班级|程序|语言|教材|目标|方式|手段", value):
                return value, min(confidence, 0.25)
            if len(value) < 2:
                return value, 0.0
            if re.fullmatch(r"[\u4e00-\u9fff·]{2,5}", value):
                return value, max(confidence, 0.82)
            return value, min(confidence, 0.55)

        if any(k in name for k in ("题目", "标题", "主题")):
            if re.search(r"第\s*\d+\s*章|[一二三四五六七八九十]+、", value):
                return value, max(confidence, 0.82)
        return value, confidence

    def add(name: str, value: str, reason: str, confidence: float):
        value = str(value or "").strip(" \t\r\n:：,，;；")
        value, confidence = adjust_candidate(name, value, reason, confidence)
        if not value or value == name or value in field_names:
            return
        if confidence <= 0:
            return
        bucket = candidates.setdefault(name, [])
        if any(x["value"] == value for x in bucket):
            return
        bucket.append({"value": value[:120], "reason": reason, "confidence_hint": round(confidence, 2)})

    # Text-level patterns: "字段：值" and simple adjacent cells/lines.
    plain = text or ""
    non_empty_tokens = [t.strip() for t in re.split(r"[\t\n\r]+", plain) if t.strip()]
    for f in fields:
        name = f["name"]
        for m in re.finditer(rf"{re.escape(name)}[ \t]*[:：][ \t]*([^\n\r\t<]{{1,80}})", plain):
            add(name, m.group(1), "字段名后冒号内容", 0.86)
        if not ocr_lines:
            for i, token in enumerate(non_empty_tokens):
                if token == name and i + 1 < len(non_empty_tokens):
                    add(name, non_empty_tokens[i + 1], "字段名单元格后的相邻文本", 0.72)

    # Geometry-level patterns: same row right side and vertically nearby values.
    for f in fields:
        name = f["name"]
        label_lines = [ln for ln in ocr_lines if name in ln["text"]]
        for label in label_lines:
            tail = label["text"].split(name, 1)[-1].strip(" :：-—")
            if tail:
                add(name, tail, f"{label['id']} 字段名同框后续文本", 0.78)

            right = [
                ln for ln in ocr_lines
                if ln is not label and same_row(label, ln) and ln["box"][0] >= label["box"][2] - 2
            ]
            right.sort(key=lambda ln: (abs(ln["cy"] - label["cy"]), ln["box"][0]))
            for ln in right[:4]:
                add(name, ln["text"], f"{label['id']} 同行右侧 {ln['id']}", 0.78)

            below = [
                ln for ln in ocr_lines
                if ln is not label and ln["box"][1] >= label["box"][3] - 2
                and abs(ln["cx"] - label["cx"]) <= max(80, (label["box"][2] - label["box"][0]) * 1.4)
            ]
            below.sort(key=lambda ln: (ln["box"][1] - label["box"][3], abs(ln["cx"] - label["cx"])))
            for ln in below[:2]:
                add(name, ln["text"], f"{label['id']} 下方相邻 {ln['id']}", 0.55)

    # Existing lightweight table reconstruction is often better than raw OCR
    # order for form-like documents. Use it only as candidate generation; the
    # model still decides or marks missing.
    if ocr_lines:
        try:
            tables = detect_tables(lines or [])
        except Exception:
            tables = []
        for table in tables[:2]:
            rows = table.get("rows") or []
            for f in fields:
                name = f["name"]
                for r_idx, row in enumerate(rows):
                    for c_idx, cell in enumerate(row):
                        if name not in str(cell):
                            continue
                        # Same row, to the right, until another requested field.
                        vals = []
                        for c2 in range(c_idx + 1, len(row)):
                            val = str(row[c2] or "").strip()
                            if not val:
                                continue
                            if any(other in val for other in field_names):
                                break
                            vals.append(val)
                        if vals:
                            add(name, " ".join(vals), "表格同一行右侧文本", 0.82)

                        # Same column / next column below, useful for long cells split
                        # across several visual rows.
                        vals = []
                        for r2 in range(r_idx + 1, min(len(rows), r_idx + 7)):
                            for c2 in (c_idx, c_idx + 1):
                                if c2 >= len(rows[r2]):
                                    continue
                                val = str(rows[r2][c2] or "").strip()
                                if not val:
                                    continue
                                if any(other in val for other in field_names):
                                    break
                                vals.append(val)
                            if vals and len(" ".join(vals)) > 140:
                                break
                        if vals:
                            add(name, " ".join(vals), "表格下方连续文本", 0.70)

                        # Some local OCR table clustering places right-column values
                        # slightly above their label. Keep this as a lower-priority
                        # candidate for textbook-like cells.
                        vals = []
                        for r2 in range(max(0, r_idx - 4), r_idx):
                            for c2 in (c_idx, c_idx + 1):
                                if c2 >= len(rows[r2]):
                                    continue
                                val = str(rows[r2][c2] or "").strip()
                                if val and not any(other in val for other in field_names):
                                    vals.append(val)
                        if vals:
                            add(name, " ".join(vals), "表格上方/右侧连续文本", 0.62)

    for key in list(candidates):
        candidates[key].sort(key=lambda x: float(x.get("confidence_hint") or 0), reverse=True)
    return {k: v[:8] for k, v in candidates.items()}


def extraction_context(text: str, lines: list[dict] | None = None, fields: list[dict] | None = None, max_lines: int = 80) -> str:
    ocr_lines = compact_ocr_lines(lines)
    if not ocr_lines:
        return (text or "")[:4000]
    selected = []
    seen = set()
    field_names = [f["name"] for f in (fields or [])]

    def add_line(ln):
        if ln["id"] not in seen:
            selected.append(ln)
            seen.add(ln["id"])

    label_lines = [ln for ln in ocr_lines if any(name in ln["text"] for name in field_names)]
    for label in label_lines:
        add_line(label)
        nearby = [
            ln for ln in ocr_lines
            if ln is not label and (
                same_row(label, ln)
                or (0 <= ln["box"][1] - label["box"][3] <= 90 and abs(ln["cx"] - label["cx"]) <= 160)
            )
        ]
        nearby.sort(key=lambda ln: (abs(ln["cy"] - label["cy"]), ln["box"][0]))
        for ln in nearby[:8]:
            add_line(ln)
    if not selected:
        selected = ocr_lines[:max_lines]
    selected.sort(key=lambda x: (round(x["cy"] / 10) * 10, x["box"][0]))
    rows = []
    for ln in selected[:max_lines]:
        x0, y0, x1, y1 = ln["box"]
        rows.append(f"{ln['id']} ({x0},{y0},{x1},{y1}) {ln['text'][:100]}")
    plain = (text or "")[:1000]
    return "OCR_LINES_WITH_BOXES:\n" + "\n".join(rows) + "\n\nPLAIN_TEXT:\n" + plain


def field_prompt(text: str, fields: list[dict], lines: list[dict] | None = None) -> list[dict]:
    field_names = [f["name"] for f in fields]
    candidates = field_candidates(text, fields, lines)
    system = (
        "你是严谨的文档字段抽取器。只基于提供的 OCR 文本、行坐标和候选值抽取字段。"
        "不要猜测，不要补全不存在的内容。只返回一个 JSON 对象，不要 Markdown，不要代码块，不要解释。"
    )
    user = (
        "请直接输出 JSON，不要输出思考过程。\n"
        "字段名：\n"
        + "\n".join(f"- {name}" for name in field_names)
        + "\n\n候选值（优先从这里选择；候选不可靠时可以判定 missing）：\n"
        + json.dumps(candidates, ensure_ascii=False)
        + "\n\n返回格式必须是：\n"
        + '{"fields":[{"name":"字段名","value":"抽取值","confidence":0.86,"evidence":"包含抽取值的原文片段","status":"ok"}]}'
        + "\n\n规则："
        + "\n1. name 必须与字段名完全一致。"
        + "\n2. value 必须来自 OCR 原文或候选值，不能改写、不能猜。"
        + "\n3. evidence 必须复制包含 value 的原文片段；不要写“原文证据或空字符串”这类占位文案。"
        + "\n4. 找不到字段或空单元格时 value/evidence 为空，confidence=0，status=missing。"
        + "\n5. 表格字段优先看字段名同一行右侧、同一行后续单元格、下方相邻单元格。"
        + "\n6. 如果候选值本身是另一个字段名，不能作为结果。"
        + "\n\nOCR 上下文：\n"
        + extraction_context(text, lines, fields)
    )
    return [
        {"role": "system", "content": system},
        {"role": "user", "content": user}
    ]


def fallback_field_extract(text: str, fields: list[dict], extractor: str, lines: list[dict] | None = None) -> tuple[list[dict], str]:
    results = []
    joined = "\n".join(ln.strip() for ln in (text or "").splitlines() if ln.strip())
    candidates = field_candidates(text, fields, lines)
    for f in fields:
        name = f["name"]
        value = ""
        evidence = ""
        confidence = 0.0
        if candidates.get(name):
            top = candidates[name][0]
            hint = float(top.get("confidence_hint") or 0.0)
            if hint >= 0.50:
                value = top["value"]
                evidence = f"{top['reason']}: {value}"
                confidence = hint
        m = re.search(rf"{re.escape(name)}[ \t]*[:：][ \t]*([^\n\r,，;；]{{1,80}})", joined)
        if m and not value:
            value = m.group(1).strip()
            evidence = m.group(0).strip()
            confidence = 0.72
        status = "ok" if confidence >= FIELD_CONF_THRESHOLD else ("low_confidence" if value else "missing")
        results.append({"name": name, "value": value, "confidence": round(confidence, 2), "evidence": evidence, "status": status})
    return results, extractor


def normalize_field_results(obj, fields: list[dict], extractor: str) -> tuple[list[dict], str]:
    raw = obj.get("fields") if isinstance(obj, dict) else obj
    by_name = {}
    if isinstance(raw, list):
        for item in raw:
            if isinstance(item, dict):
                name = str(item.get("name") or "").strip()
                if name:
                    by_name[name] = item
    results = []
    for f in fields:
        name = f["name"]
        item = by_name.get(name, {})
        value = str(item.get("value") or "").strip()
        evidence = str(item.get("evidence") or "").strip()
        try:
            confidence = float(item.get("confidence") if item.get("confidence") is not None else 0)
        except Exception:
            confidence = 0.0
        confidence = max(0.0, min(1.0, confidence))
        status = str(item.get("status") or "").strip()
        if status not in ("ok", "low_confidence", "missing"):
            status = ""
        if value and confidence < FIELD_MIN_VALUE_CONFIDENCE:
            value = ""
            evidence = ""
            status = "missing"
            confidence = 0.0
        if not status:
            status = "ok" if confidence >= FIELD_CONF_THRESHOLD and value else ("low_confidence" if value else "missing")
        if confidence < FIELD_CONF_THRESHOLD and status == "ok":
            status = "low_confidence"
        results.append({"name": name, "value": value, "confidence": round(confidence, 2), "evidence": evidence, "status": status})
    return results, extractor


def validate_field_results(text: str, fields: list[dict], results: list[dict], lines: list[dict] | None = None) -> list[dict]:
    fallback, _ = fallback_field_extract(text, fields, "regex", lines)
    by_name = {item["name"]: item for item in fallback}
    checked = []
    for item in results:
        fb = by_name.get(item["name"])
        evidence = item.get("evidence") or ""
        value = item.get("value") or ""
        evidence_ok = bool(evidence) and evidence in (text or "")
        value_ok = bool(value) and value in (text or "")
        if fb and fb.get("value") and (not value or (not evidence_ok and not value_ok)):
            item = dict(item)
            item["value"] = fb["value"]
            item["confidence"] = max(float(item.get("confidence") or 0), fb["confidence"])
            item["evidence"] = fb["evidence"]
            item["status"] = "ok" if item["confidence"] >= FIELD_CONF_THRESHOLD else "low_confidence"
        elif value_ok and not evidence_ok:
            item = dict(item)
            item["evidence"] = value
        checked.append(item)
    return checked


def local_qwen_assets_status() -> dict:
    return {
        "model": LOCAL_QWEN_MODEL,
        "url": LOCAL_QWEN_URL,
        "model_path": str(LOCAL_QWEN_MODEL_PATH),
        "model_exists": LOCAL_QWEN_MODEL_PATH.exists(),
        "llama_server_path": str(LLAMA_SERVER_EXE),
        "llama_server_exists": LLAMA_SERVER_EXE.exists(),
    }


def ensure_local_qwen_sidecar():
    global _qwen_sidecar_process
    if requests is None:
        return
    try:
        probe = requests.get(LOCAL_QWEN_URL.rsplit("/", 2)[0] + "/models", timeout=1)
        if probe.status_code < 500:
            return
    except Exception:
        pass
    if not LLAMA_SERVER_EXE.exists() or not LOCAL_QWEN_MODEL_PATH.exists():
        return
    if _qwen_sidecar_process and _qwen_sidecar_process.poll() is None:
        return
    args = [
        str(LLAMA_SERVER_EXE),
        "-m", str(LOCAL_QWEN_MODEL_PATH),
        "--host", "127.0.0.1",
        "--port", "11435",
        "-c", "4096",
        "-ngl", "0",
    ]
    creationflags = subprocess.CREATE_NO_WINDOW if os.name == "nt" else 0
    _qwen_sidecar_process = subprocess.Popen(
        args,
        cwd=str(ROOT),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        creationflags=creationflags,
    )
    deadline = time.time() + 8
    while time.time() < deadline:
        try:
            requests.get(LOCAL_QWEN_URL.rsplit("/", 2)[0] + "/models", timeout=1)
            return
        except Exception:
            time.sleep(0.5)


def extract_fields_local(text: str, fields: list[dict], lines: list[dict] | None = None) -> tuple[list[dict], str]:
    if not fields:
        return [], ""
    if requests is None:
        return fallback_field_extract(text, fields, "local-qwen-unavailable", lines)
    ensure_local_qwen_sidecar()
    payload = {"model": LOCAL_QWEN_REQUEST_MODEL, "messages": field_prompt(text, fields, lines), "temperature": 0.2, "top_p": 0.8, "max_tokens": 1200}
    try:
        r = requests.post(LOCAL_QWEN_URL, json=payload, timeout=90)
        r.raise_for_status()
        content = r.json()["choices"][0]["message"]["content"]
        results, extractor = normalize_field_results(parse_json_object(content), fields, f"local-{LOCAL_QWEN_MODEL}")
        return validate_field_results(text, fields, results, lines), extractor
    except Exception:
        return fallback_field_extract(text, fields, "local-qwen-unavailable", lines)


def extract_fields_cloud(text: str, fields: list[dict], lines: list[dict] | None = None) -> tuple[list[dict], str]:
    if not fields:
        return [], ""
    key = get_secret("DASHSCOPE_API_KEY")
    if not key or requests is None:
        return fallback_field_extract(text, fields, "cloud-qwen-unavailable", lines)
    payload = {"model": "qwen-plus", "messages": field_prompt(text, fields, lines), "temperature": 0, "max_tokens": 1200}
    try:
        r = requests.post(DASHSCOPE_CHAT_URL, headers={"Authorization": f"Bearer {key}", "Content-Type": "application/json"}, json=payload, timeout=90)
        r.raise_for_status()
        content = r.json()["choices"][0]["message"]["content"]
        results, extractor = normalize_field_results(parse_json_object(content), fields, "dashscope-qwen-plus")
        return validate_field_results(text, fields, results, lines), extractor
    except Exception:
        return fallback_field_extract(text, fields, "cloud-qwen-unavailable", lines)


def run_paddle_vl(data: bytes, filename: str) -> list[dict]:
    token = get_secret("PADDLEOCR_VL_TOKEN")
    if not token or requests is None:
        raise RuntimeError("PaddleOCR-VL token is not configured")
    headers = {"Authorization": f"bearer {token}"}
    optional_payload = {"useDocOrientationClassify": False, "useDocUnwarping": False, "useChartRecognition": False}
    with io.BytesIO(data) as f:
        files = {"file": (safe_name(filename), f)}
        form = {"model": PADDLE_VL_MODEL, "optionalPayload": json.dumps(optional_payload)}
        job_response = requests.post(PADDLE_VL_JOB_URL, headers=headers, data=form, files=files, timeout=120)
    job_response.raise_for_status()
    job_id = job_response.json()["data"]["jobId"]
    jsonl_url = ""
    deadline = time.time() + 180
    while time.time() < deadline:
        res = requests.get(f"{PADDLE_VL_JOB_URL}/{job_id}", headers=headers, timeout=30)
        res.raise_for_status()
        data = res.json()["data"]
        state = data.get("state")
        if state == "done":
            jsonl_url = data["resultUrl"]["jsonUrl"]
            break
        if state == "failed":
            raise RuntimeError(data.get("errorMsg") or "PaddleOCR-VL failed")
        time.sleep(3)
    if not jsonl_url:
        raise RuntimeError("PaddleOCR-VL timed out")
    jsonl = requests.get(jsonl_url, timeout=60)
    jsonl.raise_for_status()
    pages = []
    for line in jsonl.text.strip().splitlines():
        if not line.strip():
            continue
        result = json.loads(line)["result"]
        for res in result.get("layoutParsingResults", []):
            markdown = res.get("markdown") or {}
            raw_html = markdown.get("text", "") or ""
            safe_html = sanitize_result_html(raw_html)
            table_pairs = html_table_key_value_text(raw_html)
            plain_text = html_to_plain_text(raw_html) if safe_html else raw_html
            if table_pairs:
                plain_text = table_pairs + "\n" + plain_text
            pruned = res.get("prunedResult") or {}
            layout_boxes = ((pruned.get("layout_det_res") or {}).get("boxes") or [])
            parsing = pruned.get("parsing_res_list") or []
            pages.append({
                "text": plain_text.strip(),
                "html": safe_html,
                "markdown": raw_html,
                "output_images": res.get("outputImages") or {},
                "layout_boxes": layout_boxes,
                "parsing": parsing,
                "width": pruned.get("width"),
                "height": pruned.get("height"),
            })
    return pages


def insert_field_results(conn, task_id: str, page_id: str, fields: list[dict]):
    for item in fields:
        conn.execute("""INSERT INTO field_results
            (id, task_id, page_id, field_name, value, confidence, evidence, status)
            VALUES (?,?,?,?,?,?,?,?)""",
            (uuid.uuid4().hex[:12], task_id, page_id, item["name"], item.get("value", ""),
             float(item.get("confidence") or 0), item.get("evidence", ""), item.get("status", "")))


def fetch_page_fields(conn, page_id: str) -> list[dict]:
    rows = conn.execute("""SELECT field_name,value,confidence,evidence,status
                           FROM field_results WHERE page_id=? ORDER BY rowid""", (page_id,)).fetchall()
    return [{
        "name": r["field_name"],
        "value": r["value"] or "",
        "confidence": float(r["confidence"] or 0),
        "evidence": r["evidence"] or "",
        "status": r["status"] or "",
    } for r in rows]


def page_dict(conn, row) -> dict:
    d = dict(row)
    d["lines"] = json.loads(d.pop("lines_json") or "[]")
    d["fields"] = fetch_page_fields(conn, d["id"])
    d["image_url"] = "/" + d["image_path"] if d.get("image_path") else None
    d["thumb_url"] = "/" + d["thumb_path"] if d.get("thumb_path") else None
    d["annotated_url"] = "/" + d["annotated_path"] if d.get("annotated_path") else None
    return d


def task_dict(conn, row, include_pages=False) -> dict:
    d = dict(row)
    d["fields"] = json.loads(d.pop("fields_json") or "[]")
    d["source_names"] = json.loads(d.pop("source_names_json") or "[]")
    d["has_fields"] = bool(d["has_fields"])
    if include_pages:
        rows = conn.execute("SELECT * FROM pages WHERE task_id=? ORDER BY page_index", (d["id"],)).fetchall()
        d["pages"] = [page_dict(conn, r) for r in rows]
    return d


def task_export_data(task_id: str):
    with db() as c:
        task = c.execute("SELECT * FROM tasks WHERE id=?", (task_id,)).fetchone()
        if not task:
            raise HTTPException(404, "Task not found")
        return task_dict(c, task, include_pages=True)

# ── FastAPI ──────────────────────────────────────────────────────────────────
@asynccontextmanager
async def lifespan(app: FastAPI):
    init_db()
    print("Loading models...")
    get_models()
    print(f"Ready · backend: {current_backend()}")
    yield


app = FastAPI(title="PP-OCRv6 Local Studio", lifespan=lifespan)

app.mount("/static", StaticFiles(directory=str(STATIC)), name="static")
app.mount("/data",   StaticFiles(directory=str(DATA)),   name="data")

@app.get("/favicon.ico")
async def favicon():
    return Response(status_code=204)

@app.get("/", response_class=HTMLResponse)
async def index():
    return (STATIC / "index.html").read_text(encoding="utf-8")

@app.get("/tutorial", response_class=HTMLResponse)
async def tutorial():
    return (STATIC / "tutorial.html").read_text(encoding="utf-8")

@app.get("/info")
async def info():
    avail = ort.get_available_providers()
    cfg = load_config()
    return {
        "backend":  current_backend(),
        "model_variant": current_model_variant(),
        "providers_available": avail,
        "models_available": available_model_variants(),
        "config":   cfg,
        "version":  "0.3.1",
        "model": {
            "name":     MODEL_VARIANTS.get(current_model_variant(), MODEL_VARIANTS["tiny"])["label"],
            "official_size": MODEL_VARIANTS.get(current_model_variant(), MODEL_VARIANTS["tiny"])["official_size"],
            "dict_size": 6906,
        },
        "workbench": {
            "max_pages": MAX_TASK_PAGES,
            "max_upload_mb": MAX_UPLOAD_BYTES // 1024 // 1024,
            "field_confidence_threshold": FIELD_CONF_THRESHOLD,
            "local_text_model": LOCAL_QWEN_MODEL,
            "local_qwen": local_qwen_assets_status(),
            "cloud_text_model": "qwen-plus",
            "cloud_vl_model": PADDLE_VL_MODEL,
        },
        "usage": {
            "trial_limit_pages": TRIAL_PAGE_LIMIT,
            "trial_used_pages": get_usage_pages(),
            "trial_remaining_pages": max(0, TRIAL_PAGE_LIMIT - get_usage_pages()),
        }
    }

@app.get("/health")
async def health():
    return {"status": "ok", "backend": current_backend()}


@app.get("/usage")
async def usage_get():
    used = get_usage_pages()
    return {
        "trial_limit_pages": TRIAL_PAGE_LIMIT,
        "trial_used_pages": used,
        "trial_remaining_pages": max(0, TRIAL_PAGE_LIMIT - used),
        "payment_reserved": True,
    }


@app.post("/tasks")
async def tasks_create(
    files: list[UploadFile] = File(...),
    template: UploadFile | None = File(None),
    fields_json: str = Form("[]"),
    mode: str = Form("local"),
):
    mode = mode if mode in ("local", "cloud") else "local"
    uploaded = []
    for f in files:
        data = await f.read()
        if data:
            uploaded.append((f.filename or "upload", data))
    if not uploaded:
        raise HTTPException(400, "请选择 PDF 或图片文件")

    if template is not None and template.filename:
        fields = fields_from_excel(await template.read())
    else:
        fields = normalize_fields(fields_json)

    pages = build_page_inputs(uploaded)
    ensure_trial_available(len(pages))

    task_id = uuid.uuid4().hex[:12]
    source_names = [safe_name(name) for name, _ in uploaded]
    title = source_names[0] if len(source_names) == 1 else f"{source_names[0]} 等 {len(source_names)} 个文件"
    now = _now()
    total_bytes = sum(len(data) for _, data in uploaded)

    cloud_pages = []
    cloud_error = ""
    if mode == "cloud":
        for name, data in uploaded:
            try:
                cloud_pages.extend(run_paddle_vl(data, name))
            except Exception as exc:
                cloud_error = "云端解析暂不可用，已回退到本地 OCR"
                cloud_pages = []
                break

    with db() as c:
        c.execute("""INSERT INTO tasks
            (id, created_at, updated_at, title, mode, status, total_pages,
             processed_pages, failed_pages, source_names_json, fields_json,
             has_fields, source_bytes, trial_pages_used, error)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (task_id, now, now, title, mode, "running", len(pages), 0, 0,
             json.dumps(source_names, ensure_ascii=False),
             json.dumps(fields, ensure_ascii=False), 1 if fields else 0,
             total_bytes, len(pages), cloud_error))
        c.commit()

    processed = failed = 0
    used_before = get_usage_pages()
    set_usage_pages(used_before + len(pages))

    for idx, page in enumerate(pages):
        page_id = uuid.uuid4().hex[:12]
        img = page["image"]
        image_path = TASK_PAGES / f"{page_id}.jpg"
        thumb_path = THUMBS / f"{page_id}.jpg"
        annotated_path = ANNOTATED / f"{page_id}.jpg"
        img.save(image_path, "JPEG", quality=90)
        make_thumb(img, thumb_path)
        lines = []
        text = ""
        page_html = ""
        page_markdown = ""
        det_ms = rec_ms = total_ms = 0
        backend = current_backend()
        extractor = ""
        page_status = "done"
        page_error = ""
        try:
            if mode == "cloud" and idx < len(cloud_pages) and (cloud_pages[idx].get("text") or cloud_pages[idx].get("html")):
                cloud_page = cloud_pages[idx]
                text = (cloud_page.get("text") or html_to_plain_text(cloud_page.get("html") or "")).strip()
                page_html = cloud_page.get("html") or ""
                page_markdown = cloud_page.get("markdown") or ""
                backend = PADDLE_VL_MODEL
                out_images = cloud_page.get("output_images") or {}
                if not download_cloud_image(out_images.get("layout_det_res") or "", annotated_path):
                    img.thumbnail((1600, 1600), Image.LANCZOS)
                    img.save(annotated_path, "JPEG", quality=85)
            else:
                lines, det_ms, rec_ms = run_ocr(img, load_config())
                total_ms = round(det_ms + rec_ms)
                text = "\n".join(l["text"] for l in lines)
                annotate_image(img, lines, annotated_path)

            if fields:
                field_results, extractor = (
                    extract_fields_cloud(text, fields, lines) if mode == "cloud"
                    else extract_fields_local(text, fields, lines)
                )
            else:
                field_results = []
            processed += 1
        except Exception as exc:
            page_status = "failed"
            page_error = str(exc)
            field_results = []
            failed += 1

        with db() as c:
            c.execute("""INSERT INTO pages
                (id, task_id, page_index, source_name, page_number, status,
                 text, lines_json, backend, det_ms, rec_ms, total_ms,
                 image_path, thumb_path, annotated_path, extractor, error, html, markdown)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                (page_id, task_id, idx, page["source_name"], page["page_number"], page_status,
                 text, json.dumps(lines, ensure_ascii=False), backend,
                 round(det_ms), round(rec_ms), round(total_ms),
                 public_path(image_path), public_path(thumb_path), public_path(annotated_path),
                 extractor, page_error, page_html, page_markdown))
            insert_field_results(c, task_id, page_id, field_results)
            c.commit()

    status = "done" if failed == 0 else ("failed" if processed == 0 else "partial")
    with db() as c:
        c.execute("""UPDATE tasks SET status=?, updated_at=?, processed_pages=?, failed_pages=?
                     WHERE id=?""", (status, _now(), processed, failed, task_id))
        c.commit()
        row = c.execute("SELECT * FROM tasks WHERE id=?", (task_id,)).fetchone()
        return task_dict(c, row, include_pages=True)


@app.post("/templates/fields")
async def template_fields(template: UploadFile = File(...)):
    data = await template.read()
    fields = fields_from_excel(data)
    return {"fields": fields}


@app.get("/tasks")
async def tasks_list(limit: int = 100, offset: int = 0, q: str = ""):
    with db() as c:
        if q:
            rows = c.execute("""SELECT * FROM tasks
                                WHERE title LIKE ? OR source_names_json LIKE ?
                                ORDER BY created_at DESC LIMIT ? OFFSET ?""",
                             (f"%{q}%", f"%{q}%", limit, offset)).fetchall()
            total = c.execute("""SELECT COUNT(*) FROM tasks
                                 WHERE title LIKE ? OR source_names_json LIKE ?""",
                              (f"%{q}%", f"%{q}%")).fetchone()[0]
        else:
            rows = c.execute("SELECT * FROM tasks ORDER BY created_at DESC LIMIT ? OFFSET ?",
                             (limit, offset)).fetchall()
            total = c.execute("SELECT COUNT(*) FROM tasks").fetchone()[0]
        return {"total": total, "items": [task_dict(c, r, include_pages=False) for r in rows]}


@app.get("/tasks/{task_id}")
async def tasks_detail(task_id: str):
    with db() as c:
        row = c.execute("SELECT * FROM tasks WHERE id=?", (task_id,)).fetchone()
        if not row:
            raise HTTPException(404, "Task not found")
        return task_dict(c, row, include_pages=True)


@app.get("/tasks/{task_id}/pages")
async def tasks_pages(task_id: str):
    with db() as c:
        rows = c.execute("SELECT * FROM pages WHERE task_id=? ORDER BY page_index", (task_id,)).fetchall()
        return {"items": [page_dict(c, r) for r in rows]}


def task_plain_text(task: dict) -> str:
    parts = []
    for p in task.get("pages", []):
        head = f"## {p['source_name']} 第 {p['page_number']} 页"
        parts.append(head)
        parts.append(p.get("text") or "")
    return "\n\n".join(parts).strip()


def task_fields_matrix(task: dict):
    field_names = [f["name"] for f in task.get("fields", [])]
    rows = []
    for p in task.get("pages", []):
        vals = {f["name"]: f.get("value", "") for f in p.get("fields", [])}
        rows.append([vals.get(name, "") for name in field_names])
    return field_names, rows


@app.get("/tasks/{task_id}/export")
async def tasks_export(task_id: str, fmt: str = Query("xlsx", pattern="^(xlsx|txt|md|csv|html)$")):
    task = task_export_data(task_id)
    name = safe_name(task.get("title") or "task").rsplit(".", 1)[0]
    has_fields = bool(task.get("fields"))
    if fmt == "xlsx":
        if not has_fields:
            raise HTTPException(400, "未指定字段名时不支持导出 Excel")
        if Workbook is None:
            raise HTTPException(500, "openpyxl is not installed; cannot export xlsx")
        headers, rows = task_fields_matrix(task)
        wb = Workbook()
        ws = wb.active
        ws.title = "字段结果"
        ws.append(headers)
        for row in rows:
            ws.append(row)
        header_fill = PatternFill("solid", fgColor="EEF2FF")
        header_font = Font(bold=True, color="111827")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
        for idx, header in enumerate(headers, start=1):
            ws.column_dimensions[get_column_letter(idx)].width = min(max(12, len(str(header)) * 2), 36)
        ws.freeze_panes = "A2"
        buf = io.BytesIO()
        wb.save(buf)
        return Response(
            content=buf.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": content_disposition(f"{name}.xlsx")}
        )

    text = task_plain_text(task)
    if fmt == "txt":
        return PlainTextResponse(text, headers={"Content-Disposition": content_disposition(f"{name}.txt")})
    if fmt == "md":
        return PlainTextResponse(text, headers={"Content-Disposition": content_disposition(f"{name}.md")})
    if fmt == "csv":
        if has_fields:
            headers, rows = task_fields_matrix(task)
            buf = io.StringIO()
            writer = csv.writer(buf)
            writer.writerow(headers)
            writer.writerows(rows)
            body = buf.getvalue()
        else:
            body = text
        return Response(content=body, media_type="text/csv; charset=utf-8",
                        headers={"Content-Disposition": content_disposition(f"{name}.csv")})
    if fmt == "html":
        if has_fields:
            headers, rows = task_fields_matrix(task)
            trs = ["<tr>" + "".join(f"<th>{html.escape(str(h))}</th>" for h in headers) + "</tr>"]
            for row in rows:
                trs.append("<tr>" + "".join(f"<td>{html.escape(str(v))}</td>" for v in row) + "</tr>")
            body = "<table>" + "\n".join(trs) + "</table>"
        else:
            body = f"<pre>{html.escape(text)}</pre>"
        return HTMLResponse(f"<!doctype html><meta charset='utf-8'><body>{body}</body>",
                            headers={"Content-Disposition": content_disposition(f"{name}.html")})

@app.post("/ocr")
async def ocr_endpoint(file: UploadFile = File(...),
                       save: bool = Query(True)):
    if not file.content_type or not file.content_type.startswith("image/"):
        raise HTTPException(400, "Please upload an image file")
    data = await file.read()
    try:
        img = Image.open(io.BytesIO(data)).convert("RGB")
    except Exception:
        raise HTTPException(400, "Cannot decode image")
    if max(img.size) > 6000:
        img.thumbnail((4096, 4096), Image.LANCZOS)

    cfg = load_config()
    rec_id = uuid.uuid4().hex[:12]
    lines, det_ms, rec_ms = run_ocr(img, cfg)
    tables = detect_tables(lines)

    # Save assets if history enabled
    upload_path = thumb_path = annotated_path = ""
    if save and cfg.get("save_history", True):
        ext = (file.filename or "upload.jpg").split(".")[-1].lower()
        if ext not in ("jpg", "jpeg", "png", "webp", "bmp"):
            ext = "jpg"
        upload_path = f"data/uploads/{rec_id}.{ext}"
        thumb_path  = f"data/thumbs/{rec_id}.jpg"
        annotated_path = f"data/annotated/{rec_id}.jpg"
        with open(DATA / "uploads" / f"{rec_id}.{ext}", "wb") as f:
            f.write(data)
        make_thumb(img, DATA / "thumbs" / f"{rec_id}.jpg")

    annotated_b64 = annotate_image(img, lines,
        save_path=(DATA / "annotated" / f"{rec_id}.jpg") if upload_path else None)

    if save and cfg.get("save_history", True):
        with db() as c:
            c.execute("""INSERT INTO history
                (id, created_at, filename, n_boxes, det_ms, rec_ms, total_ms,
                 backend, text, lines_json, tables_json,
                 upload_path, thumb_path, annotated_path)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                (rec_id, time.time(), file.filename or "untitled",
                 len(lines), round(det_ms), round(rec_ms), round(det_ms+rec_ms),
                 current_backend(),
                 "\n".join(l["text"] for l in lines),
                 json.dumps(lines, ensure_ascii=False),
                 json.dumps(tables, ensure_ascii=False),
                 upload_path, thumb_path, annotated_path))
            c.commit()

    return JSONResponse({
        "id": rec_id,
        "filename": file.filename,
        "lines":    lines,
        "tables":   tables,
        "text":     "\n".join(l["text"] for l in lines),
        "det_ms":   round(det_ms),
        "rec_ms":   round(rec_ms),
        "total_ms": round(det_ms + rec_ms),
        "n_boxes":  len(lines),
        "n_tables": len(tables),
        "backend":  current_backend(),
        "model_variant": current_model_variant(),
        "annotated": annotated_b64,
        "thumb_url": "/" + thumb_path if thumb_path else None,
    })

# ── History ──────────────────────────────────────────────────────────────────
@app.get("/history")
async def history_list(limit: int = 50, offset: int = 0, q: str = ""):
    with db() as c:
        if q:
            rows = c.execute(
                """SELECT id,created_at,filename,n_boxes,total_ms,backend,thumb_path,tables_json
                   FROM history WHERE filename LIKE ? OR text LIKE ?
                   ORDER BY created_at DESC LIMIT ? OFFSET ?""",
                (f"%{q}%", f"%{q}%", limit, offset)).fetchall()
            total = c.execute(
                "SELECT COUNT(*) FROM history WHERE filename LIKE ? OR text LIKE ?",
                (f"%{q}%", f"%{q}%")).fetchone()[0]
        else:
            rows = c.execute(
                """SELECT id,created_at,filename,n_boxes,total_ms,backend,thumb_path,tables_json
                   FROM history ORDER BY created_at DESC LIMIT ? OFFSET ?""",
                (limit, offset)).fetchall()
            total = c.execute("SELECT COUNT(*) FROM history").fetchone()[0]
    items = []
    for r in rows:
        item = dict(r)
        try:
            item["n_tables"] = len(json.loads(item.pop("tables_json") or "[]"))
        except Exception:
            item["n_tables"] = 0
        items.append(item)
    return {"total": total, "items": items}

@app.get("/history/{rec_id}")
async def history_detail(rec_id: str):
    with db() as c:
        row = c.execute("SELECT * FROM history WHERE id=?", (rec_id,)).fetchone()
        if not row:
            raise HTTPException(404, "Not found")
        d = dict(row)
        d["lines"]  = json.loads(d.pop("lines_json")  or "[]")
        d["tables"] = json.loads(d.pop("tables_json") or "[]")
        return d


@app.get("/history/{rec_id}/export")
async def history_export(rec_id: str, fmt: str = Query("csv", pattern="^(csv|xlsx|md|html|json|txt)$")):
    with db() as c:
        row = c.execute("SELECT * FROM history WHERE id=?", (rec_id,)).fetchone()
        if not row:
            raise HTTPException(404, "Not found")
    tables = json.loads(row["tables_json"] or "[]")
    text   = row["text"] or ""
    filename = (row["filename"] or "ocr").rsplit(".", 1)[0]

    if fmt == "txt":
        return PlainTextResponse(text, headers={
            "Content-Disposition": content_disposition(f"{filename}.txt")})
    if fmt == "json":
        return JSONResponse(dict(row), headers={
            "Content-Disposition": content_disposition(f"{filename}.json")})
    if fmt == "csv":
        body = tables_to_csv(tables) if tables else text
        return Response(content=body, media_type="text/csv; charset=utf-8",
                        headers={"Content-Disposition": content_disposition(f"{filename}.csv")})
    if fmt == "xlsx":
        if not tables:
            tables = [{"rows": [[text]], "n_rows": 1, "n_cols": 1}]
        body = tables_to_xlsx(tables)
        return Response(content=body,
                        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        headers={"Content-Disposition": content_disposition(f"{filename}.xlsx")})
    if fmt == "md":
        body = tables_to_markdown(tables) if tables else text
        return PlainTextResponse(body, headers={
            "Content-Disposition": content_disposition(f"{filename}.md")})
    if fmt == "html":
        body = tables_to_html(tables) if tables else f"<pre>{text}</pre>"
        return HTMLResponse(body, headers={
            "Content-Disposition": content_disposition(f"{filename}.html")})

@app.delete("/history/{rec_id}")
async def history_delete(rec_id: str):
    with db() as c:
        row = c.execute("SELECT upload_path,thumb_path,annotated_path FROM history WHERE id=?",
                        (rec_id,)).fetchone()
        if row:
            for p in (row["upload_path"], row["thumb_path"], row["annotated_path"]):
                if p:
                    fp = ROOT / "webapp" / p
                    try: fp.unlink()
                    except Exception: pass
        c.execute("DELETE FROM history WHERE id=?", (rec_id,))
        c.commit()
    return {"ok": True}

@app.post("/export")
async def export_inline(payload: dict = Body(...),
                        fmt: str = Query("csv", pattern="^(csv|xlsx|md|html|txt)$")):
    """Stateless export — payload: {tables: [...], text: '...'}"""
    tables = payload.get("tables", [])
    text   = payload.get("text", "")
    name   = (payload.get("filename") or "ocr").rsplit(".",1)[0]
    if fmt == "txt":
        return PlainTextResponse(text, headers={
            "Content-Disposition": content_disposition(f"{name}.txt")})
    if fmt == "csv":
        body = tables_to_csv(tables) if tables else text
        return Response(content=body, media_type="text/csv; charset=utf-8",
                        headers={"Content-Disposition": content_disposition(f"{name}.csv")})
    if fmt == "xlsx":
        if not tables:
            tables = [{"rows": [[text]], "n_rows": 1, "n_cols": 1}]
        body = tables_to_xlsx(tables)
        return Response(content=body,
                        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        headers={"Content-Disposition": content_disposition(f"{name}.xlsx")})
    if fmt == "md":
        body = tables_to_markdown(tables) if tables else text
        return PlainTextResponse(body, headers={
            "Content-Disposition": content_disposition(f"{name}.md")})
    if fmt == "html":
        body = tables_to_html(tables) if tables else f"<pre>{text}</pre>"
        return HTMLResponse(body, headers={
            "Content-Disposition": content_disposition(f"{name}.html")})


@app.post("/history/clear")
async def history_clear():
    with db() as c:
        c.execute("DELETE FROM history")
        c.commit()
    for d in (UPLOADS, ANNOTATED, THUMBS):
        for f in d.iterdir():
            try: f.unlink()
            except Exception: pass
    return {"ok": True}

# ── Settings ─────────────────────────────────────────────────────────────────
@app.get("/settings")
async def settings_get():
    return load_config()

@app.put("/settings")
async def settings_put(payload: dict = Body(...)):
    global _det_sess, _rec_sess, _current_provider, _current_model_variant
    cfg = save_config(payload)
    # Reload models if provider/threads changed
    _det_sess = _rec_sess = _current_provider = _current_model_variant = None
    get_models()
    return {"ok": True, "config": cfg, "backend": current_backend()}

@app.post("/settings/reset")
async def settings_reset():
    cfg = save_config(dict(DEFAULT_CFG))
    return {"ok": True, "config": cfg}

if __name__ == "__main__":
    port = int(os.environ.get("DOC_WORKBENCH_PORT", os.environ.get("PORT", "8766")))
    print("PP-OCRv6 Local Studio")
    print(f"http://localhost:{port}")
    uvicorn.run(app, host="127.0.0.1", port=port, log_level="warning")
